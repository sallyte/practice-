# file1 = open("test.txt","w",encoding="UTF-8")

# file1.write('hello world')

# print('이설희', file = file1)

# file1.close()




from openpyxl import Workbook

# 새로운 워크북 생성
wb = Workbook()

# 기본적으로 생성된 시트를 활성화
ws = wb.active

# 셀 데이터를 읽어서 출력하기
for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=2):
    for cell in row:
        print(cell.value, end=" ")  # 셀의 값을 출력
    print()  # 행 끝에 줄바꿈


# # 셀에 데이터 쓰기
# ws['A1'] = "Name"

# ws['B1'] = "Age"
# ws['A2'] = "John Doe"
# ws['B2'] = 25
# ws['A3'] = "Jane Doe"
# ws['B3'] = 30

# 엑셀 파일로 저장
wb.save("sample.xlsx")



